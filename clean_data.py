"""Asset Management data normalizer v2.

The old cleaner flattened different snapshots into six files. That caused
foreign-key breaks, duplicate assets, multiline serial numbers, and loss of
provenance. This version builds a relational staging dataset from raw Data/:

- one canonical row per employee, location, model, and physical asset;
- one asset unit per quantity/serial number;
- stable keys: source asset ID first, then serial number, then observation key;
- assignments, transfers, maintenance, warehouse stock, network data and
  source observations kept as separate linked tables;
- unresolved/conflicting records are not silently guessed: they go to
  data_quality_issues.csv and are marked needs_review.

Raw files under Data/ are never modified. Outputs are written to
Data clear/v2_normalized/.
"""

from __future__ import annotations

import csv
import hashlib
import json
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
RAW_DIR = BASE_DIR / "Data"
OUT_DIR = BASE_DIR / "Data clear" / "v2_normalized"


def text(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).replace("\xa0", " ").replace("\r", "\n")
    return " ".join(s.split()).strip()


def code(value: Any) -> str:
    return re.sub(r"[^A-Z0-9_\-]", "", text(value).upper().replace(" ", ""))


def ascii_key(value: Any) -> str:
    s = unicodedata.normalize("NFKD", text(value))
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", " ", s.lower()).strip()


def slug(value: Any, fallback: str = "UNKNOWN") -> str:
    s = ascii_key(value).upper().replace(" ", "_")
    s = re.sub(r"[^A-Z0-9_\-]", "", s)
    return s[:50] or fallback


def digest(*parts: Any, size: int = 10) -> str:
    raw = "|".join(text(p) for p in parts)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:size].upper()


def serials(value: Any) -> list[str]:
    s = text(value)
    if not s or s.lower() in {"nan", "none", "n/a", "-", "null"}:
        return []
    return [code(x) for x in re.split(r"[\n,;]+", s) if code(x)]


def parse_int(value: Any, default: int = 1) -> int:
    try:
        return max(0, int(float(text(value))))
    except (TypeError, ValueError):
        return default


def source_ref(path: Path, sheet: str, row: int) -> str:
    try:
        rel = path.relative_to(RAW_DIR).as_posix()
    except ValueError:
        rel = path.relative_to(BASE_DIR).as_posix()
    return f"{rel}#{sheet}!row={row}"


def source_id(ref: str) -> str:
    return f"SRC-{digest(ref, size=12)}"


def read_sheet(path: Path, sheet: str | None = None, data_only: bool = True) -> list[list[Any]]:
    wb = load_workbook(path, read_only=True, data_only=data_only)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.worksheets[0]
    return [list(row) for row in ws.iter_rows(values_only=True)]


def find_header(rows: list[list[Any]], required: Iterable[str], limit: int = 12) -> int | None:
    required = {ascii_key(x) for x in required}
    for i, row in enumerate(rows[:limit]):
        values = {ascii_key(x) for x in row if text(x)}
        if required.issubset(values):
            return i
    return None


def dict_rows(path: Path, sheet: str | None = None, required: Iterable[str] = ()) -> list[dict[str, Any]]:
    rows = read_sheet(path, sheet)
    header_i = find_header(rows, required) if required else 0
    if header_i is None:
        return []
    header = [text(x) for x in rows[header_i]]
    result: list[dict[str, Any]] = []
    for row_i, row in enumerate(rows[header_i + 1 :], header_i + 2):
        if not any(text(x) for x in row):
            continue
        result.append({header[j]: row[j] if j < len(row) else None for j in range(len(header)) if header[j]})
        result[-1]["__row__"] = row_i
    return result


def csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def first_value(row: dict[str, Any], *names: str) -> Any:
    normalized = {ascii_key(k): v for k, v in row.items()}
    for name in names:
        value = normalized.get(ascii_key(name))
        if value is not None and text(value):
            return value
    return None


def write_csv(name: str, rows: list[dict[str, Any]], fields: list[str]) -> None:
    path = OUT_DIR / name
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") if row.get(field, "") is not None else "" for field in fields})


def append_note(old: str, new: str) -> str:
    values = [x for x in [text(old), text(new)] if x]
    return " | ".join(dict.fromkeys(values))


class Normalizer:
    def __init__(self) -> None:
        self.issues: list[dict[str, Any]] = []
        self.issue_seq = 0
        self.locations: dict[str, dict[str, Any]] = {}
        self.networks: dict[str, dict[str, Any]] = {}
        self.departments: dict[str, dict[str, Any]] = {}
        self.employees: dict[str, dict[str, Any]] = {}
        self.employee_by_name: dict[str, str] = {}
        self.categories: dict[str, dict[str, Any]] = {}
        self.models: dict[str, dict[str, Any]] = {}
        self.assets: dict[str, dict[str, Any]] = {}
        self.asset_by_id: dict[str, str] = {}
        self.asset_by_serial: dict[str, str] = {}
        self.observations: list[dict[str, Any]] = []
        self.assignments: dict[str, dict[str, Any]] = {}
        self.transfers: list[dict[str, Any]] = []
        self.maintenance: list[dict[str, Any]] = []
        self.warehouses: dict[str, dict[str, Any]] = {}
        self.stock_transactions: list[dict[str, Any]] = []
        self.stock_balances: list[dict[str, Any]] = []
        self.planned_shipments: list[dict[str, Any]] = []
        self.relationships: list[dict[str, Any]] = []
        self.manifest: list[dict[str, Any]] = []

    def issue(self, kind: str, message: str, ref: str = "", entity_type: str = "", entity_key: str = "", **extra: Any) -> None:
        self.issue_seq += 1
        row = {
            "issue_id": f"ISSUE-{self.issue_seq:05d}",
            "severity": extra.pop("severity", "REVIEW"),
            "issue_type": kind,
            "message": message,
            "source_ref": ref,
            "entity_type": entity_type,
            "entity_key": entity_key,
            **extra,
        }
        self.issues.append(row)

    def manifest_file(self, path: Path, purpose: str, priority: int, row_count: int = 0) -> None:
        try:
            source_file = path.relative_to(RAW_DIR).as_posix()
        except ValueError:
            source_file = path.relative_to(BASE_DIR).as_posix()
        self.manifest.append({"source_file": source_file, "purpose": purpose, "priority": priority, "rows_read": row_count})

    def ensure_location(self, location_code: str, name: str = "", brand: str = "", region: str = "", ref: str = "", inferred: bool = False) -> str:
        location_code = code(location_code)
        if not location_code:
            location_code = "HCM"
        if location_code not in self.locations:
            self.locations[location_code] = {
                "location_code": location_code,
                "name": text(name) or location_code,
                "brand": text(brand),
                "region": text(region),
                "open_date": "",
                "store_manager": "",
                "phone": "",
                "area_manager": "",
                "organization_code": "MAP-VN",
                "source_refs": ref,
                "inferred": str(bool(inferred)).lower(),
            }
        else:
            row = self.locations[location_code]
            for field, value in [("name", name), ("brand", brand), ("region", region)]:
                if text(value) and (not text(row.get(field)) or row.get(field) == location_code):
                    row[field] = text(value)
            row["source_refs"] = append_note(row.get("source_refs", ""), ref)
            if inferred:
                row["inferred"] = "true"
        return location_code

    def ensure_category(self, category_code: str, name: str) -> str:
        category_code = code(category_code) or "OFF"
        if category_code not in self.categories:
            self.categories[category_code] = {"category_code": category_code, "name": text(name) or category_code, "source_refs": "raw mapping"}
        return category_code

    def category_for(self, raw_category: Any = "", hardware: Any = "", name: Any = "") -> str:
        raw = ascii_key(raw_category)
        hw = ascii_key(hardware)
        combined = f"{raw} {hw} {ascii_key(name)}"
        if "network" in raw or any(x in combined for x in ["firewall", "switch", "access point", "wifi", "router"]):
            return self.ensure_category("NET", "Network & Security")
        if "pdt" in raw or "pdt" in combined:
            return self.ensure_category("PDT", "PDT / Handheld Scanner")
        if "back end" in raw or any(x in combined for x in ["printer", "tablet", "server"]):
            return self.ensure_category("BACKEND", "Back-end / Store Support")
        if "pos" in raw or any(x in combined for x in ["cash drawer", "bill printer", "pole display", "scanner", "mini pc", "desktop"]):
            return self.ensure_category("POS", "Point of Sale")
        if any(x in combined for x in ["monitor", "display"]):
            return self.ensure_category("MON", "Monitor & Display")
        if any(x in combined for x in ["printer", "epson", "canon", "sato", "postek"]):
            return self.ensure_category("PRN", "Printer")
        if any(x in combined for x in ["laptop", "notebook", "thinkpad", "latitude", "elitebook", "probook"]):
            return self.ensure_category("LAP", "Laptop / Computer")
        return self.ensure_category("OFF", "Office Equipment")

    def ensure_model(self, model_name: Any, category_code: str, manufacturer: str = "", ref: str = "") -> str:
        display = text(model_name) or "Unspecified equipment"
        key = f"{ascii_key(display)}|{category_code}"
        model_code = f"M-{digest(key, size=10)}"
        if model_code not in self.models:
            self.models[model_code] = {
                "model_code": model_code,
                "name": display,
                "manufacturer": text(manufacturer),
                "category_code": category_code,
                "source_refs": ref,
            }
        else:
            self.models[model_code]["source_refs"] = append_note(self.models[model_code].get("source_refs", ""), ref)
        return model_code

    def register_store(self, store_code: Any, name: Any = "", brand: Any = "", region: Any = "", open_date: Any = "", manager: Any = "", phone: Any = "", area_manager: Any = "", ref: str = "", inferred: bool = False) -> str:
        location_code = self.ensure_location(store_code, name, brand, region, ref, inferred)
        row = self.locations[location_code]
        for field, value in [("open_date", open_date), ("store_manager", manager), ("phone", phone), ("area_manager", area_manager)]:
            if text(value) and not text(row.get(field)):
                row[field] = text(value)
        return location_code

    def register_asset(self, *, asset_id: Any = "", serial_number: Any = "", name: Any = "", category_code: str, model_code: str, location_code: str = "HCM", intended_location_code: str = "", employee_id: str = "", department_code: str = "", status: str = "AVAILABLE", notes: str = "", ref: str = "", source_kind: str = "", needs_review: bool = False, raw_qty: int = 1, unit_index: int = 1, extra: dict[str, Any] | None = None) -> str:
        asset_id = code(asset_id)
        serial_number = code(serial_number)
        # Resolve against either identity before creating a new canonical row.
        # This is critical when one source has Asset ID and another only has S/N.
        key = self.asset_by_id.get(asset_id, "") if asset_id else ""
        if not key and serial_number:
            key = self.asset_by_serial.get(serial_number, "")
        if not key:
            if asset_id:
                key = f"AID:{asset_id}"
            elif serial_number:
                key = f"SN:{serial_number}"
            else:
                key = f"OBS:{digest(source_kind, ref, unit_index, name)}"
        location_code = self.ensure_location(location_code or "HCM", ref=ref, inferred=True)
        intended_location_code = code(intended_location_code)
        if intended_location_code:
            self.ensure_location(intended_location_code, ref=ref, inferred=True)
        existing = self.assets.get(key)
        if existing:
            if serial_number and existing.get("serial_number") and serial_number != existing["serial_number"]:
                needs_review = True
                self.issue("ASSET_KEY_CONFLICT", f"Asset key {key} received two serial numbers", ref, "asset", key)
            if location_code != existing.get("location_code"):
                needs_review = True
                self.issue("DUPLICATE_SERIAL_CROSS_LOCATION", f"Same physical key observed at {existing.get('location_code')} and {location_code}", ref, "asset", key)
            for field, value in [("asset_id", asset_id), ("serial_number", serial_number), ("name", name), ("employee_id", employee_id), ("department_code", department_code), ("intended_location_code", intended_location_code)]:
                if text(value) and not text(existing.get(field)):
                    existing[field] = text(value)
            existing["notes"] = append_note(existing.get("notes", ""), notes)
            existing["source_refs"] = append_note(existing.get("source_refs", ""), ref)
            existing["needs_review"] = str(bool(existing.get("needs_review") == "true" or needs_review)).lower()
            if status == "UNDER_MAINTENANCE":
                existing["status"] = status
            elif existing.get("status") == "AVAILABLE" and status in {"IN_USE", "RESERVED"}:
                existing["status"] = status
        else:
            asset_code = asset_id or (f"SN-{serial_number}" if serial_number else f"DATA-{digest(key, size=12)}")
            self.assets[key] = {
                "asset_key": key,
                "asset_code": asset_code,
                "asset_id": asset_id,
                "serial_number": serial_number,
                "name": text(name) or "Unspecified equipment",
                "category_code": category_code,
                "model_code": model_code,
                "location_code": location_code,
                "intended_location_code": intended_location_code,
                "department_code": department_code,
                "employee_id": employee_id,
                "status": status,
                "raw_qty": raw_qty,
                "unit_index": unit_index,
                "notes": text(notes),
                "source_kind": source_kind,
                "source_refs": ref,
                "needs_review": str(bool(needs_review)).lower(),
                **(extra or {}),
            }
        self.observations.append({
            "observation_id": f"OBS-{digest(key, ref, unit_index, size=12)}",
            "asset_key": key,
            "observation_status": status,
            "observed_location_code": location_code,
            "intended_location_code": intended_location_code,
            "source_kind": source_kind,
            "source_ref": ref,
            "raw_asset_id": asset_id,
            "raw_serial_number": serial_number,
            "raw_name": text(name),
            "raw_qty": raw_qty,
            "unit_index": unit_index,
        })
        if asset_id:
            previous = self.asset_by_id.get(asset_id)
            if previous and previous != key:
                self.issue("DUPLICATE_ASSET_ID", f"Asset ID {asset_id} maps to multiple canonical keys", ref, "asset", key)
            self.asset_by_id[asset_id] = key
        if serial_number:
            previous = self.asset_by_serial.get(serial_number)
            if previous and previous != key:
                self.issue("DUPLICATE_SERIAL", f"Serial number {serial_number} maps to multiple canonical keys", ref, "asset", key)
            self.asset_by_serial[serial_number] = key
        if employee_id:
            assignment_key = f"ASN:{key}:{employee_id}"
            self.assignments[assignment_key] = {
                "assignment_key": assignment_key,
                "asset_key": key,
                "employee_id": employee_id,
                "assigned_date": "",
                "actual_return_date": "",
                "status": "ACTIVE",
                "assigned_by": "DATA_IMPORT",
                "source_ref": ref,
                "notes": "Derived from source holder/Employee Id",
            }
        return key

    def find_asset(self, asset_id: Any = "", serial_number: Any = "") -> str:
        aid = code(asset_id)
        sn = code(serial_number)
        return self.asset_by_id.get(aid) or self.asset_by_serial.get(sn) or ""

    def add_relationship(self, source_type: str, source_key: str, relation: str, target_type: str, target_key: str, ref: str = "") -> None:
        if source_key and target_key:
            self.relationships.append({"source_type": source_type, "source_key": source_key, "relation": relation, "target_type": target_type, "target_key": target_key, "source_ref": ref})

    def build_locations(self) -> None:
        checklist = RAW_DIR / "Audit 2026 device" / "Data store" / "Check List.xlsx"
        if checklist.exists():
            for sheet in load_workbook(checklist, read_only=True, data_only=True).sheetnames:
                rows = read_sheet(checklist, sheet)
                for row_i, row in enumerate(rows, 1):
                    values = [text(x) for x in row]
                    for idx, value in enumerate(values):
                        if not re.fullmatch(r"[A-Za-z]{1,3}\d{2,3}", value or ""):
                            continue
                        name = values[idx + 1] if idx + 1 < len(values) else ""
                        if name and ascii_key(name) in {"store code", "am", "region"}:
                            continue
                        ref = source_ref(checklist, sheet, row_i)
                        self.register_store(value, name=name, ref=ref)
            self.manifest_file(checklist, "store master/checklist", 1)
        firewall = RAW_DIR / "Audit 2026 device" / "Data store" / "Firewall Serial Number.xlsx"
        if firewall.exists():
            rows = dict_rows(firewall, required=["Storecode", "Serial Number"])
            for row in rows:
                sc = code(first_value(row, "Storecode", "Store code"))
                if not sc:
                    continue
                ref = source_ref(firewall, "Sheet1", int(row["__row__"]))
                self.register_store(sc, ref=ref, inferred=True)
                self.networks[sc] = {
                    "location_code": sc,
                    "firewall_serial": code(first_value(row, "Serial Number")),
                    "ip_gateway": text(first_value(row, "IP Gateway")),
                    "ip_wan": text(first_value(row, "IP WAN")),
                    "ip_wan_default_gateway": text(first_value(row, "IP WAN Default gateway")),
                    "note": text(first_value(row, "Note")),
                    "source_ref": ref,
                }
            self.manifest_file(firewall, "store firewall/network", 1, len(rows))
        # Store codes are authoritative in per-store audit files; register all of them.
        inventory_dirs = [RAW_DIR / "Audit 2026 device" / "Data store" / "Audit", RAW_DIR / "Audit 2026 device" / "Data store" / "!!!"]
        for directory in inventory_dirs:
            for path in sorted(directory.glob("*.xlsx")) if directory.exists() else []:
                sc = code(path.stem)
                if sc:
                    self.register_store(sc, name="", ref=path.relative_to(RAW_DIR).as_posix(), inferred=True)
                    self.manifest_file(path, "store asset audit", 1)
        details = RAW_DIR / "Asset management" / "Store" / "Details"
        for path in sorted(details.glob("*.xlsx")) if details.exists() else []:
            sc = code(path.stem)
            if sc:
                self.register_store(sc, ref=path.relative_to(RAW_DIR).as_posix(), inferred=True)
                self.manifest_file(path, "store asset detail fallback", 2)
        dashboard = RAW_DIR / "Asset management" / "Store" / "Dashboard.xlsx"
        if dashboard.exists():
            rows = dict_rows(dashboard, "All Stores", required=["Store", "Category", "Hardware Type", "Model", "Qty", "S/N"])
            for row in rows:
                sc = code(first_value(row, "Store"))
                if sc:
                    self.register_store(sc, ref=source_ref(dashboard, "All Stores", int(row["__row__"])), inferred=True)
            self.manifest_file(dashboard, "store consolidated dashboard", 3, len(rows))
        self.ensure_location("HCM", "HQ Hồ Chí Minh", "Office", "South", inferred=True)
        self.ensure_location("HN", "Office Hà Nội", "Office", "North", inferred=True)
        self.ensure_location("IT_STOCK", "Kho IT Stock", "Warehouse", "", inferred=True)
        self.ensure_location("DSV", "Kho DSV", "Warehouse", "", inferred=True)
        self.ensure_location("GEODIS", "Kho Geodis", "Warehouse", "", inferred=True)
        self.ensure_location("MAP_FASHION", "Map Fashion (transfer destination - needs confirmation)", "External", "", inferred=True)
        for row in self.locations.values():
            if row.get("inferred") == "true" and row.get("location_code") not in {"HCM", "HN", "IT_STOCK", "DSV", "GEODIS", "MAP_FASHION"}:
                if not text(row.get("name")) or row.get("name") == row.get("location_code"):
                    self.issue("MISSING_STORE_MASTER", f"Store {row['location_code']} has no reliable store name/master row", row.get("source_refs", ""), "location", row["location_code"])

    def build_departments_employees(self) -> None:
        path = RAW_DIR / "Asset management" / "Office" / "Employees.xlsx"
        rows = dict_rows(path, "Sheet1", required=["Employee Id", "Full Name", "Current Department"])
        self.manifest_file(path, "employee master", 1, len(rows))
        for row in rows:
            eid = code(first_value(row, "Employee Id"))
            name = text(first_value(row, "Full Name"))
            department = text(first_value(row, "Current Department"))
            cost_center = code(first_value(row, "Department Cost Center Id"))
            if not eid:
                self.issue("MISSING_EMPLOYEE_ID", f"Employee {name or 'unknown'} has no Employee Id", source_ref(path, "Sheet1", int(row["__row__"])), "employee", "")
                continue
            dept_code = cost_center or f"D-{slug(department)}"
            self.departments.setdefault(dept_code, {"department_code": dept_code, "name": department or dept_code, "cost_center_code": cost_center, "organization_code": "MAP-VN", "source_refs": source_ref(path, "Sheet1", int(row["__row__"]))})
            record = {
                "employee_id": eid,
                "staff_code": eid,
                "full_name": name,
                "email": text(first_value(row, "Official Email Id")),
                "status": "ACTIVE" if ascii_key(first_value(row, "Status")) in {"active", "working"} else "INACTIVE",
                "position": text(first_value(row, "Current Designation")),
                "department_code": dept_code,
                "department_name": department,
                "source_ref": source_ref(path, "Sheet1", int(row["__row__"])),
            }
            if eid in self.employees and self.employees[eid].get("full_name") != name:
                self.issue("DUPLICATE_EMPLOYEE_ID", f"Employee Id {eid} has conflicting names", record["source_ref"], "employee", eid)
            else:
                self.employees[eid] = {**self.employees.get(eid, {}), **{k: v for k, v in record.items() if text(v)}}
            if name:
                self.employee_by_name[ascii_key(name)] = eid

    def office_candidate_rows(self) -> list[tuple[Path, str, dict[str, Any]]]:
        paths = [
            (RAW_DIR / "Asset management" / "Office" / "Office assets.xlsx", 1),
            (RAW_DIR / "Audit 2026 device" / "Office assets.xlsx", 2),
        ]
        result: list[tuple[Path, str, dict[str, Any]]] = []
        for path, priority in paths:
            if not path.exists():
                continue
            sheets = load_workbook(path, read_only=True, data_only=True).sheetnames
            sheet = "Asset" if "Asset" in sheets else "Sheet1"
            rows = dict_rows(path, sheet, required=["Status", "Asset Name", "S/N", "Asset ID"])
            self.manifest_file(path, "office asset snapshot", priority, len(rows))
            for row in rows:
                result.append((path, sheet, row))
        return result

    def build_office_assets(self) -> None:
        candidates = self.office_candidate_rows()
        for path, sheet, row in candidates:
            ref = source_ref(path, sheet, int(row["__row__"]))
            aid = code(first_value(row, "Asset ID"))
            sn = code(first_value(row, "S/N", "Serial Number"))
            name = text(first_value(row, "Asset Name", "Model")) or "Office Asset"
            holder_id = code(first_value(row, "Employee Id"))
            holder_name = text(first_value(row, "Full Name"))
            if not holder_id and holder_name:
                holder_id = self.employee_by_name.get(ascii_key(holder_name), "")
            raw_dept = text(first_value(row, "Current Department"))
            dept_code = ""
            for key, value in self.departments.items():
                if ascii_key(value.get("name")) == ascii_key(raw_dept):
                    dept_code = key
                    break
            if holder_id and holder_id not in self.employees:
                self.issue("UNKNOWN_EMPLOYEE_REFERENCE", f"Office asset references Employee Id {holder_id} not in employee master", ref, "asset", aid or sn, employee_id=holder_id, holder_name=holder_name)
                holder_id = ""
            status_raw = ascii_key(first_value(row, "Status"))
            status = "IN_USE" if "use" in status_raw else "AVAILABLE"
            host = text(first_value(row, "Hostname", "Computer Name"))
            location = "HCM"
            for loc in self.locations:
                if loc not in {"HCM", "HN", "IT_STOCK", "DSV", "GEODIS", "MAP_FASHION"} and re.search(rf"(?<![A-Z0-9]){re.escape(loc)}(?:[-_]|$)", host.upper()):
                    location = loc
                    break
            category = self.category_for(name=name)
            model = self.ensure_model(name, category, ref=ref)
            notes = "; ".join(x for x in [f"Hostname: {host}" if host else "", f"CPU: {text(first_value(row, 'CPU'))}" if text(first_value(row, 'CPU')) else "", f"RAM: {text(first_value(row, 'RAM'))}" if text(first_value(row, 'RAM')) else "", f"Disk: {text(first_value(row, 'Hard Disk'))}" if text(first_value(row, 'Hard Disk')) else ""] if x)
            key = self.register_asset(asset_id=aid, serial_number=sn, name=name, category_code=category, model_code=model, location_code=location, employee_id=holder_id, department_code=dept_code, status=status, notes=notes, ref=ref, source_kind="OFFICE_ASSET", needs_review=bool((not aid and not sn) or (holder_name and not holder_id)))
            if holder_name and not holder_id:
                self.assets[key]["notes"] = append_note(self.assets[key].get("notes", ""), f"Unresolved holder name: {holder_name}")

    def choose_store_sources(self) -> dict[str, list[Path]]:
        selected: dict[str, list[Path]] = defaultdict(list)
        dirs = [RAW_DIR / "Audit 2026 device" / "Data store" / "Audit", RAW_DIR / "Audit 2026 device" / "Data store" / "!!!"]
        for directory in dirs:
            if not directory.exists():
                continue
            for path in sorted(directory.glob("*.xlsx")):
                selected[code(path.stem)].append(path)
        details = RAW_DIR / "Asset management" / "Store" / "Details"
        if details.exists():
            for path in sorted(details.glob("*.xlsx")):
                sc = code(path.stem)
                if sc and sc not in selected:
                    selected[sc].append(path)
        return selected

    def inventory_rows(self, path: Path) -> list[tuple[str, dict[str, Any]]]:
        result: list[tuple[str, dict[str, Any]]] = []
        wb = load_workbook(path, read_only=True, data_only=True)
        seen: set[str] = set()
        for sheet in wb.sheetnames:
            rows = read_sheet(path, sheet)
            header_i = find_header(rows, ["Category", "Model"])
            if header_i is None:
                continue
            header = [text(x) for x in rows[header_i]]
            normalized = {ascii_key(x): i for i, x in enumerate(header) if x}
            if "hardware type" not in normalized and "type device" not in normalized:
                continue
            for row_i, row in enumerate(rows[header_i + 1 :], header_i + 2):
                def get(*names: str) -> Any:
                    for name in names:
                        idx = normalized.get(ascii_key(name))
                        if idx is not None and idx < len(row) and text(row[idx]):
                            return row[idx]
                    return None
                if not any(text(x) for x in row):
                    continue
                raw = {"Category": get("Category"), "Hardware Type": get("Hardware Type", "Type Device", "Hardware"), "Model": get("Model"), "Qty": get("Qty", "Qty ", "SL"), "S/N": get("S/N", "Serial Number", "Serial number"), "Note": get("Note", "Remark")}
                if not any(text(raw[x]) for x in ["Hardware Type", "Model", "S/N"]):
                    continue
                fingerprint = "|".join([ascii_key(raw[x]) for x in ["Category", "Hardware Type", "Model", "Qty", "S/N", "Note"]])
                if fingerprint in seen:
                    self.issue("DUPLICATE_SOURCE_SHEET_ROW", "Same inventory row repeated in another sheet of the workbook", source_ref(path, sheet, row_i), "source", "")
                    continue
                seen.add(fingerprint)
                raw["__row__"] = row_i
                result.append((sheet, raw))
        return result

    def build_store_assets(self) -> None:
        selected = self.choose_store_sources()
        for store_code, paths in sorted(selected.items()):
            if not store_code:
                continue
            # The first source is the 2026 audit. Details is fallback only.
            path = paths[0]
            rows = self.inventory_rows(path)
            self.manifest_file(path, "selected store inventory", 1 if "Audit 2026" in path.as_posix() else 2, len(rows))
            for sheet, row in rows:
                ref = source_ref(path, sheet, int(row["__row__"]))
                cat_raw = text(row.get("Category"))
                hw = text(row.get("Hardware Type"))
                model_name = text(row.get("Model")) or hw or "Store Equipment"
                qty = parse_int(row.get("Qty"), 1)
                found_serials = serials(row.get("S/N"))
                if len(found_serials) > qty:
                    self.issue("SERIAL_COUNT_EXCEEDS_QTY", f"Serial count {len(found_serials)} exceeds quantity {qty}; expanded quantity to preserve units", ref, "asset", "", quantity=qty, serial_count=len(found_serials))
                    qty = len(found_serials)
                if qty <= 0:
                    self.issue("INVALID_QUANTITY", "Quantity is zero/invalid; skipped row", ref, "source", "")
                    continue
                category = self.category_for(cat_raw, hw, model_name)
                model = self.ensure_model(model_name, category, ref=ref)
                note = text(row.get("Note"))
                at_warehouse = any(x in ascii_key(note) for x in ["dsv", "geodis", "warehouse", "kho"])
                physical_location = "DSV" if "dsv" in ascii_key(note) else ("GEODIS" if "geodis" in ascii_key(note) else store_code)
                status = "AVAILABLE" if at_warehouse else "IN_USE"
                if "new" in ascii_key(note) and not at_warehouse:
                    status = "RESERVED"
                for unit in range(1, qty + 1):
                    sn = found_serials[unit - 1] if unit <= len(found_serials) else ""
                    if not sn:
                        self.issue("MISSING_SERIAL_UNIT", f"Unit {unit}/{qty} has no serial number", ref, "asset", f"STORE:{store_code}:{int(row['__row__'])}:{unit}", store_code=store_code, quantity=qty)
                    key = self.register_asset(serial_number=sn, name=hw or model_name, category_code=category, model_code=model, location_code=physical_location, intended_location_code=store_code if at_warehouse else "", status=status, notes=append_note(note, f"Source quantity {qty}; unit {unit}"), ref=ref, source_kind="STORE_AUDIT", needs_review=not bool(sn), raw_qty=qty, unit_index=unit, extra={"source_store_code": store_code})
                    self.add_relationship("asset", key, "OBSERVED_AT", "location", physical_location, ref)
                    if at_warehouse:
                        self.add_relationship("asset", key, "INTENDED_FOR", "location", store_code, ref)

    def build_warehouses_and_stock(self) -> None:
        dsv = RAW_DIR / "Asset management" / "Warehouse" / "Asset_List_DSV.xlsx"
        if dsv.exists():
            self.warehouses["DSV"] = {"warehouse_code": "DSV", "name": "Kho DSV", "location_code": "DSV", "source_ref": dsv.relative_to(RAW_DIR).as_posix()}
            rows = dict_rows(dsv, "Asset List", required=["Broad Category", "Type", "SL", "S/N"])
            self.manifest_file(dsv, "DSV warehouse physical inventory", 1, len(rows))
            for row in rows:
                ref = source_ref(dsv, "Asset List", int(row["__row__"]))
                qty = parse_int(first_value(row, "SL"), 1)
                sns = serials(first_value(row, "S/N"))
                qty = max(qty, len(sns), 1)
                category = self.category_for(first_value(row, "Broad Category"), first_value(row, "Type"), first_value(row, "Type"))
                model = self.ensure_model(first_value(row, "Type", "Capital Installation / Expenditure (VND)"), category, ref=ref)
                for unit in range(1, qty + 1):
                    self.register_asset(serial_number=sns[unit - 1] if unit <= len(sns) else "", name=first_value(row, "Type") or "DSV Stock", category_code=category, model_code=model, location_code="DSV", status="AVAILABLE", ref=ref, source_kind="WAREHOUSE_DSV", needs_review=unit > len(sns), raw_qty=qty, unit_index=unit)
        geodis = RAW_DIR / "Asset management" / "Warehouse" / "Geodis device.xlsx"
        if geodis.exists():
            self.warehouses["GEODIS"] = {"warehouse_code": "GEODIS", "name": "Kho Geodis", "location_code": "GEODIS", "source_ref": geodis.relative_to(RAW_DIR).as_posix()}
            rows = dict_rows(geodis, "Sheet1", required=["Model", "Asset ID", "S/N"])
            self.manifest_file(geodis, "Geodis warehouse devices", 1, len(rows))
            for row in rows:
                ref = source_ref(geodis, "Sheet1", int(row["__row__"]))
                category = self.category_for(name=first_value(row, "Model"))
                model = self.ensure_model(first_value(row, "Model"), category, ref=ref)
                notes = "; ".join(x for x in [f"Hostname: {text(first_value(row, 'Hostname'))}" if text(first_value(row, 'Hostname')) else "", f"Email: {text(first_value(row, 'Mail'))}" if text(first_value(row, 'Mail')) else "", f"CPU: {text(first_value(row, 'CPU'))}" if text(first_value(row, 'CPU')) else "", f"RAM: {text(first_value(row, 'Ram'))}" if text(first_value(row, 'Ram')) else "", f"Disk: {text(first_value(row, 'Hard Disk'))}" if text(first_value(row, 'Hard Disk')) else ""] if x)
                self.register_asset(asset_id=first_value(row, "Asset ID"), serial_number=first_value(row, "S/N"), name=first_value(row, "Model"), category_code=category, model_code=model, location_code="GEODIS", status="AVAILABLE", notes=notes, ref=ref, source_kind="WAREHOUSE_GEODIS")
        transitions = RAW_DIR / "Asset management" / "Transititons" / "IT tool.xlsx"
        if transitions.exists():
            stock_rows = dict_rows(transitions, "Infra Device (stock) ", required=["Item", "Type Device", "Stock"])
            self.manifest_file(transitions, "IT stock balance", 1, len(stock_rows))
            for row in stock_rows:
                self.stock_balances.append({"item": text(first_value(row, "Item")), "device_type": text(first_value(row, "Type Device")), "reported_stock": text(first_value(row, "Stock")), "note": text(first_value(row, "Note")), "source_ref": source_ref(transitions, "Infra Device (stock) ", int(row["__row__"]))})
            tx_rows = dict_rows(transitions, "Infra Device (transactions)", required=["Item", "Transaction Type", "Quantity"])
            self.manifest_file(transitions, "IT stock transactions", 1, len(tx_rows))
            for row in tx_rows:
                ref = source_ref(transitions, "Infra Device (transactions)", int(row["__row__"]))
                tx_type = code(first_value(row, "Transaction Type"))
                tx_type = "IN" if tx_type == "IN" else ("OUT" if tx_type == "OUT" else "ADJUSTMENT")
                self.stock_transactions.append({"transaction_key": f"STK-{digest(ref)}", "item": text(first_value(row, "Item")), "transaction_type": tx_type, "quantity": parse_int(first_value(row, "Quantity"), 1), "transaction_date": text(first_value(row, "Date")), "unit": text(first_value(row, "Unit")), "assigned_to": text(first_value(row, "Assigned")), "serial_numbers": ";".join(serials(first_value(row, "S/N"))), "description": text(first_value(row, "Description")), "remark": text(first_value(row, "Remark")), "source_ref": ref})
            mini_rows = dict_rows(transitions, "Mini PC status", required=["Asset Code", "Model", "Serial number"])
            self.manifest_file(transitions, "mini PC status", 1, len(mini_rows))
            for row in mini_rows:
                ref = source_ref(transitions, "Mini PC status", int(row["__row__"]))
                if not code(first_value(row, "Serial number")) and not code(first_value(row, "Asset Code")):
                    continue
                category = self.category_for(name=first_value(row, "Model"))
                model = self.ensure_model(first_value(row, "Model"), category, manufacturer=first_value(row, "Vendor"), ref=ref)
                self.register_asset(asset_id=f"MINIPC-{code(first_value(row, 'Asset Code'))}", serial_number=first_value(row, "Serial number"), name=first_value(row, "Model") or "Mini PC", category_code=category, model_code=model, location_code="IT_STOCK", status="AVAILABLE", notes=append_note(text(first_value(row, "Note")), f"MAC LAN: {text(first_value(row, 'MAC Address LAN'))}"), ref=ref, source_kind="MINIPC_STATUS", needs_review=not bool(code(first_value(row, "Serial number"))))

    def build_broken_and_transfers(self) -> None:
        path = BASE_DIR / "Data clear" / "05_Broken_and_Transfers_Cleaned.xlsx"
        if not path.exists():
            path = RAW_DIR / "Asset management" / "Office" / "Broken 0 sửa.xlsx"
        if path.exists():
            for sheet in load_workbook(path, read_only=True, data_only=True).sheetnames:
                rows = dict_rows(path, sheet, required=["Asset Name", "S/N"])
                if not rows:
                    continue
                self.manifest_file(path, "broken/transfer source", 1, len(rows))
                if ascii_key(sheet).startswith("office broken") or path.name.startswith("Broken"):
                    for row in rows:
                        ref = source_ref(path, sheet, int(row["__row__"]))
                        aid = first_value(row, "Asset ID")
                        sn = first_value(row, "S/N")
                        key = self.find_asset(aid, sn)
                        if not key:
                            category = self.category_for(name=first_value(row, "Asset Name"))
                            model = self.ensure_model(first_value(row, "Asset Name"), category, ref=ref)
                            key = self.register_asset(asset_id=aid, serial_number=sn, name=first_value(row, "Asset Name") or "Broken Asset", category_code=category, model_code=model, location_code="HCM", status="UNDER_MAINTENANCE", notes=first_value(row, "Details"), ref=ref, source_kind="BROKEN_ASSET", needs_review=True)
                            self.issue("BROKEN_ASSET_NOT_IN_INVENTORY", "Broken asset could not be matched to office/store inventory; created review record", ref, "asset", key)
                        else:
                            self.assets[key]["status"] = "UNDER_MAINTENANCE"
                            self.assets[key]["notes"] = append_note(self.assets[key].get("notes", ""), first_value(row, "Details"))
                        ticket = {
                            "ticket_key": f"TKT-{digest(key, ref)}",
                            "ticket_no": f"DATA-BRK-{digest(key, ref, size=8)}",
                            "asset_key": key,
                            "title": f"Hư hỏng: {text(first_value(row, 'Asset Name')) or 'Asset'}",
                            "description": text(first_value(row, "Details")) or "Broken asset reported in source",
                            "priority": "HIGH",
                            "status": "OPEN",
                            "reported_by_employee_id": "",
                            "source_ref": ref,
                            "needs_review": "true",
                        }
                        self.maintenance.append(ticket)
                        self.issue("MISSING_MAINTENANCE_REPORTER", "Source has no employee reporter; import may use controlled system admin fallback", ref, "maintenance", ticket["ticket_key"])
                elif ascii_key(sheet).startswith("transfer"):
                    for row in rows:
                        ref = source_ref(path, sheet, int(row["__row__"]))
                        aid = first_value(row, "Asset ID")
                        sn = first_value(row, "S/N")
                        key = self.find_asset(aid, sn)
                        if not key:
                            category = self.category_for(name=first_value(row, "Asset Name"))
                            model = self.ensure_model(first_value(row, "Asset Name"), category, ref=ref)
                            key = self.register_asset(asset_id=aid, serial_number=sn, name=first_value(row, "Asset Name") or "Transferred asset", category_code=category, model_code=model, location_code="HCM", status="AVAILABLE", ref=ref, source_kind="TRANSFER_ASSET", needs_review=True)
                            self.issue("TRANSFER_ASSET_NOT_IN_INVENTORY", "Transfer record created an asset pending identity confirmation", ref, "asset", key)
                        asset = self.assets[key]
                        transfer = {
                            "transfer_key": f"TRF-{digest(key, ref)}",
                            "asset_key": key,
                            "from_location_code": asset.get("location_code", "HCM"),
                            "to_location_code": "MAP_FASHION",
                            "from_department_code": asset.get("department_code", ""),
                            "to_department_code": "",
                            "reason": "Source list says transfer to Map Fashion; date/approval not provided",
                            "status": "PENDING",
                            "requested_by_employee_id": "",
                            "source_ref": ref,
                            "needs_review": "true",
                        }
                        self.transfers.append(transfer)
                        self.issue("TRANSFER_METADATA_INCOMPLETE", "Transfer destination is inferred from worksheet name; date, requester and approval are missing", ref, "transfer", transfer["transfer_key"])
        store_broken = RAW_DIR / "Asset management" / "Store" / "Broken.xlsx"
        if store_broken.exists():
            rows = read_sheet(store_broken, "Sheet1")
            if rows:
                ref = source_ref(store_broken, "Sheet1", 1)
                values = rows[0]
                sc = code(values[0]) if len(values) > 0 else ""
                category = self.category_for(values[1] if len(values) > 1 else "", values[2] if len(values) > 2 else "", values[3] if len(values) > 3 else "")
                model = self.ensure_model(values[3] if len(values) > 3 else "Broken store device", category, ref=ref)
                key = self.find_asset(serial_number=values[5] if len(values) > 5 else "")
                if not key:
                    key = self.register_asset(serial_number=values[5] if len(values) > 5 else "", name=values[3] if len(values) > 3 else "Broken store device", category_code=category, model_code=model, location_code=sc or "HCM", status="UNDER_MAINTENANCE", notes=values[6] if len(values) > 6 else "", ref=ref, source_kind="BROKEN_STORE", needs_review=True)
                    self.issue("BROKEN_STORE_NOT_IN_INVENTORY", "Store broken record was not found in audit inventory", ref, "asset", key)
                else:
                    self.assets[key]["status"] = "UNDER_MAINTENANCE"
                self.maintenance.append({"ticket_key": f"TKT-{digest(key, ref)}", "ticket_no": f"DATA-BRK-{digest(key, ref, size=8)}", "asset_key": key, "title": f"Hư hỏng: {text(values[3])}", "description": text(values[6]) if len(values) > 6 else "", "priority": "HIGH", "status": "OPEN", "reported_by_employee_id": "", "source_ref": ref, "needs_review": "true"})
                self.manifest_file(store_broken, "store broken asset", 1, 1)

    def build_planned_shipments(self) -> None:
        path = RAW_DIR / "Audit 2026 device" / "Data store" / "New Store 2026" / "Operation device for plan new store.xlsx"
        if not path.exists():
            return
        self.manifest_file(path, "planned new-store shipment, not current inventory", 2)
        wb = load_workbook(path, read_only=True, data_only=True)
        for sheet in wb.sheetnames:
            rows = self.inventory_rows(path) if sheet == wb.sheetnames[0] else []
            # inventory_rows reads all sheets and would duplicate; use direct local parse below.
            raw = read_sheet(path, sheet)
            header_i = find_header(raw, ["Category", "Model"])
            if header_i is None:
                continue
            headers = [text(x) for x in raw[header_i]]
            idx = {ascii_key(x): i for i, x in enumerate(headers) if x}
            match = re.search(r"([A-Za-z]{1,3}\d{2,3})", sheet)
            target = code(match.group(1)) if match else ""
            for row_i, row in enumerate(raw[header_i + 1 :], header_i + 2):
                if not any(text(x) for x in row):
                    continue
                def get(*names: str) -> Any:
                    for name in names:
                        j = idx.get(ascii_key(name))
                        if j is not None and j < len(row) and text(row[j]):
                            return row[j]
                    return None
                if not get("Model", "Hardware Type"):
                    continue
                ref = source_ref(path, sheet, row_i)
                self.planned_shipments.append({"shipment_key": f"SHIP-{digest(ref)}", "target_location_code": target, "category": text(get("Category")), "hardware_type": text(get("Hardware Type")), "model": text(get("Model")), "quantity": parse_int(get("Qty", "Qty "), 1), "serial_numbers": ";".join(serials(get("S/N"))), "source_ref": ref, "status": "PLANNED"})

    def validate(self) -> None:
        asset_codes = Counter(text(row.get("asset_code")) for row in self.assets.values())
        for value, count in asset_codes.items():
            if value and count > 1:
                self.issue("DUPLICATE_CANONICAL_ASSET_CODE", f"Canonical asset code {value} is not unique", "", "asset", value, severity="ERROR")
        serial_map: dict[str, list[str]] = defaultdict(list)
        for key, row in self.assets.items():
            if row.get("serial_number"):
                serial_map[row["serial_number"]].append(key)
            if row.get("location_code") not in self.locations:
                self.issue("MISSING_LOCATION_LINK", f"Asset points to missing location {row.get('location_code')}", row.get("source_refs", ""), "asset", key, severity="ERROR")
            if row.get("model_code") not in self.models:
                self.issue("MISSING_MODEL_LINK", f"Asset points to missing model {row.get('model_code')}", row.get("source_refs", ""), "asset", key, severity="ERROR")
            if row.get("category_code") not in self.categories:
                self.issue("MISSING_CATEGORY_LINK", f"Asset points to missing category {row.get('category_code')}", row.get("source_refs", ""), "asset", key, severity="ERROR")
        for sn, keys in serial_map.items():
            if len(keys) > 1:
                self.issue("DUPLICATE_CANONICAL_SERIAL", f"Serial {sn} appears in multiple canonical assets", "", "asset", ";".join(keys), severity="ERROR")
        for row in self.assignments.values():
            if row["asset_key"] not in self.assets:
                self.issue("BROKEN_ASSIGNMENT_ASSET_LINK", "Assignment points to missing asset", row.get("source_ref", ""), "assignment", row["assignment_key"], severity="ERROR")
            if row["employee_id"] not in self.employees:
                self.issue("BROKEN_ASSIGNMENT_EMPLOYEE_LINK", "Assignment points to missing employee", row.get("source_ref", ""), "assignment", row["assignment_key"], severity="ERROR")

    def write(self) -> None:
        OUT_DIR.mkdir(parents=True, exist_ok=True)
        write_csv("organizations.csv", [{"organization_code": "MAP-VN", "name": "MAP Active Vietnam", "source_refs": "normalized master"}], ["organization_code", "name", "source_refs"])
        write_csv("locations.csv", list(self.locations.values()), ["location_code", "name", "brand", "region", "open_date", "store_manager", "phone", "area_manager", "organization_code", "source_refs", "inferred"])
        write_csv("store_networks.csv", list(self.networks.values()), ["location_code", "firewall_serial", "ip_gateway", "ip_wan", "ip_wan_default_gateway", "note", "source_ref"])
        write_csv("departments.csv", list(self.departments.values()), ["department_code", "name", "cost_center_code", "organization_code", "source_refs"])
        write_csv("employees.csv", list(self.employees.values()), ["employee_id", "staff_code", "full_name", "email", "status", "position", "department_code", "department_name", "source_ref"])
        write_csv("asset_categories.csv", list(self.categories.values()), ["category_code", "name", "source_refs"])
        write_csv("asset_models.csv", list(self.models.values()), ["model_code", "name", "manufacturer", "category_code", "source_refs"])
        write_csv("assets.csv", list(self.assets.values()), ["asset_key", "asset_code", "asset_id", "serial_number", "name", "category_code", "model_code", "location_code", "intended_location_code", "department_code", "employee_id", "status", "raw_qty", "unit_index", "notes", "source_kind", "source_refs", "needs_review", "source_store_code"])
        write_csv("asset_observations.csv", self.observations, ["observation_id", "asset_key", "observation_status", "observed_location_code", "intended_location_code", "source_kind", "source_ref", "raw_asset_id", "raw_serial_number", "raw_name", "raw_qty", "unit_index"])
        write_csv("asset_assignments.csv", list(self.assignments.values()), ["assignment_key", "asset_key", "employee_id", "assigned_date", "actual_return_date", "status", "assigned_by", "source_ref", "notes"])
        write_csv("asset_transfers.csv", self.transfers, ["transfer_key", "asset_key", "from_location_code", "to_location_code", "from_department_code", "to_department_code", "reason", "status", "requested_by_employee_id", "source_ref", "needs_review"])
        write_csv("maintenance_tickets.csv", self.maintenance, ["ticket_key", "ticket_no", "asset_key", "title", "description", "priority", "status", "reported_by_employee_id", "source_ref", "needs_review"])
        write_csv("warehouses.csv", list(self.warehouses.values()), ["warehouse_code", "name", "location_code", "source_ref"])
        write_csv("stock_balances.csv", self.stock_balances, ["item", "device_type", "reported_stock", "note", "source_ref"])
        write_csv("stock_transactions.csv", self.stock_transactions, ["transaction_key", "item", "transaction_type", "quantity", "transaction_date", "unit", "assigned_to", "serial_numbers", "description", "remark", "source_ref"])
        write_csv("planned_shipments.csv", self.planned_shipments, ["shipment_key", "target_location_code", "category", "hardware_type", "model", "quantity", "serial_numbers", "source_ref", "status"])
        write_csv("relationships.csv", self.relationships, ["source_type", "source_key", "relation", "target_type", "target_key", "source_ref"])
        write_csv("source_manifest.csv", self.manifest, ["source_file", "purpose", "priority", "rows_read"])
        write_csv("data_quality_issues.csv", self.issues, ["issue_id", "severity", "issue_type", "message", "source_ref", "entity_type", "entity_key", "employee_id", "holder_name", "store_code", "quantity", "serial_count"])
        summary = {
            "version": "v2_normalized",
            "raw_dir": str(RAW_DIR),
            "output_dir": str(OUT_DIR),
            "counts": {
                "locations": len(self.locations), "departments": len(self.departments), "employees": len(self.employees), "categories": len(self.categories), "models": len(self.models), "assets": len(self.assets), "assets_with_serial": sum(bool(x.get("serial_number")) for x in self.assets.values()), "assignments": len(self.assignments), "transfers": len(self.transfers), "maintenance_tickets": len(self.maintenance), "warehouses": len(self.warehouses), "stock_transactions": len(self.stock_transactions), "planned_shipments": len(self.planned_shipments), "issues": len(self.issues), "error_issues": sum(x.get("severity") == "ERROR" for x in self.issues),
            },
            "status_counts": dict(Counter(x.get("status", "") for x in self.assets.values())),
            "store_asset_counts": dict(Counter(x.get("location_code", "") for x in self.assets.values() if x.get("source_kind") == "STORE_AUDIT")),
            "rules": ["Raw Data/ files are read-only.", "One asset unit per serial/quantity unit.", "Asset ID is preferred, then serial number, then deterministic observation key.", "Conflicts are retained in asset_observations.csv and data_quality_issues.csv.", "No production database was modified by this script."],
        }
        (OUT_DIR / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
        (OUT_DIR / "README.md").write_text(self.readme(summary), encoding="utf-8")
        print(json.dumps(summary, ensure_ascii=False, indent=2))

    def readme(self, summary: dict[str, Any]) -> str:
        return f"""# Data clear v2 — relational staging dataset\n\nGenerated from raw files under `Data/` by `clean_data.py`. Raw files are not modified.\n\n## Why v2 exists\nThe previous output mixed snapshots and aggregates. It kept quantity rows with multiline serials, duplicated office assets, and merged store/firewall rows by row position. v2 separates master data, physical assets, observations, assignments, transfers, maintenance, warehouses, and quality issues.\n\n## Link keys\n- `employee_id`: source HR Employee Id; used by `asset_assignments.csv`.\n- `department_code`: HR Department Cost Center Id where available.\n- `location_code`: normalized store code (`VA01`, `VN05`, `A001`, `VN28`, ...), plus `HCM`, `HN`, `DSV`, `GEODIS`, `IT_STOCK`.\n- `asset_key`: `AID:<Asset ID>` first, then `SN:<serial>`, then deterministic `OBS:<hash>` for units without identity.\n- `asset_code`: source Asset ID when available; otherwise deterministic generated code.\n- `model_code` and `category_code`: stable mapping tables.\n\n## Important data rules\n1. A row with `Qty=3` and three serials becomes three asset units.\n2. A row with quantity greater than serial count creates review units without serials.\n3. Duplicate serials are never silently overwritten. The canonical asset is retained, every raw observation is retained, and the conflict is listed in `data_quality_issues.csv`.\n4. Store rows marked `DSV` are physically located at `DSV` and keep the intended store in `intended_location_code`.\n5. Transfer-to-Map-Fashion records are `PENDING` and `needs_review=true` because the source has no date, requester, approval, or confirmed destination master.\n6. Planned new-store shipment data is in `planned_shipments.csv`; it is not counted as current inventory.\n\n## Output counts\n```json\n{json.dumps(summary['counts'], ensure_ascii=False, indent=2)}\n```\n\n## Import gate\nDo not import to production while `data_quality_issues.csv` contains unresolved identity/location conflicts. First review the issues, then run the importer against a staging Neon branch.\n"""


def main() -> None:
    normalizer = Normalizer()
    normalizer.build_locations()
    normalizer.build_departments_employees()
    normalizer.build_office_assets()
    normalizer.build_store_assets()
    normalizer.build_warehouses_and_stock()
    normalizer.build_broken_and_transfers()
    normalizer.build_planned_shipments()
    normalizer.validate()
    normalizer.write()


if __name__ == "__main__":
    main()
